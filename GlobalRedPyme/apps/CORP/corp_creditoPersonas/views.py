import io
import datetime
import json

import qrcode
import fitz
from cryptography.hazmat import backends
from cryptography.hazmat.primitives.serialization import pkcs12
from endesive.pdf import cms
## Libreria para agregar imagenes a pdf
from PIL import Image, ImageDraw, ImageFont
import urllib.parse

import os

from ..corp_movimientoCobros.models import Transacciones

# Establecer el directorio base del proyecto
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
from django.core.files.uploadedfile import InMemoryUploadedFile

from ...CENTRAL.central_catalogo.models import Catalogo
from .models import CreditoPersonas, CodigoCredito, RegarcarCreditos
from ...PERSONAS.personas_personas.models import Personas
from ..corp_empresas.models import Empresas
from .serializers import (
    CreditoPersonasSerializer, CreditoPersonasPersonaSerializer, CodigoCreditoSerializer,
    RegarcarCreditosSerializer,
)
# Enviar Correo
from ...config.util import sendEmail
# Importar boto3
import boto3
import tempfile
import environ
# Publicar en sns
from .producer_ifi import publish
# Consumir en sqs
from .consumer_ifi import get_queue_url
from rest_framework import status
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from django.utils import timezone
from django.conf import settings
# Swagger
from drf_yasg.utils import swagger_auto_schema
# Lectura de AWS s3
import boto3
import re
from apps.config import config
# excel
import openpyxl
# Generar codigos aleatorios
import string
import random
# Sumar minutos
from dateutil.relativedelta import relativedelta
# ObjectId
from bson import ObjectId
# logs
from ...CENTRAL.central_logs.methods import createLog, datosTipoLog, datosProductosMDP

from ...utils.extraerDatosFirmaElectronica import usuarioPropietarioFirma

# declaracion variables log
datosAux = datosProductosMDP()
datosTipoLogAux = datosTipoLog()
# asignacion datos modulo
logModulo = datosAux['modulo']
logApi = datosAux['api']
# asignacion tipo de datos
logTransaccion = datosTipoLogAux['transaccion']
logExcepcion = datosTipoLogAux['excepcion']


# CRUD
# CREAR
# 'methods' can be used to apply the same modification to multiple methods
@swagger_auto_schema(methods=['post'], request_body=CreditoPersonasSerializer)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def creditoPersonas_create(request):
    """
    Este metod sirve para crear un creditopersona
    @type request: recibe los campos de la tabla credito persona
    @rtype: DEvuelve el registro creado, caso contrario devuelve el error generado
    """
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'create/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'CREAR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)
            request.data['created_at'] = str(timezone_now)
            if 'updated_at' in request.data:
                request.data.pop('updated_at')

            if 'nombres' in request.data:
                if request.data['nombres'] != "":
                    request.data['nombresCompleto'] = f"""{request.data['nombres']} {request.data['apellidos']}"""

            tipoCredito = 'Pymes-Normales'
            if 'tipoCredito' in request.data:
                if 'Pymes-PreAprobado' in request.data['tipoCredito']:
                    tipoCredito = request.data.pop('tipoCredito')

            serializer = CreditoPersonasSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                createLog(logModel, serializer.data, logTransaccion)
                if serializer.data['estado'] == 'Nuevo' and serializer.data['tipoCredito'] == 'Pymes-Normales':
                    credito = serializer.data
                    credito['cargarOrigen'] = 'BIGPUNTOS'
                    # Publicar en la cola
                    publish(credito)
                    enviarCorreoSolicitud(request.data['email'])
                if serializer.data['estado'] == 'Nuevo' and tipoCredito == 'Pymes-PreAprobado':
                    credito = serializer.data
                    credito['tipoCredito'] = 'Pymes-PreAprobado'
                    credito['cargarOrigen'] = 'BIGPUNTOS'
                    # Publicar en la cola
                    publish(credito)
                    enviarCorreoSolicitud(request.data['email'])
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            createLog(logModel, serializer.errors, logExcepcion)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
def creditoPersonas_listOne_sinAutenticar(request, pk):
    """
    Este metod sirve para obtener un creditopersona
    @type pk: recibe el id de la tabla credito persona
    @type request: no recibe
    @rtype: DEvuelve el registro obtenido, caso contrario devuelve el error generado
    """
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'listOne/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        if request.method == 'POST':
            query = CreditoPersonas.objects.filter(pk=ObjectId(pk), tipoCredito=request.data['tipoCredito'],
                                                   state=1).first()
            if query is None:
                err = {"error": "No existe"}
                createLog(logModel, err, logExcepcion)
                return Response(err, status=status.HTTP_404_NOT_FOUND)
            # tomar el dato
            serializer = CreditoPersonasSerializer(query)
            createLog(logModel, serializer.data, logTransaccion)
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


# ENCONTRAR UNO
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def creditoPersonas_listOne(request, pk):
    """
    Este metod sirve para obtener un creditopersona
    @type pk: recibe el id de la tabla credito persona
    @type request: no recibe
    @rtype: DEvuelve el registro obtenido, caso contrario devuelve el error generado
    """
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'listOne/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            query = CreditoPersonas.objects.filter(pk=ObjectId(pk), state=1).first()
        except CreditoPersonas.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'GET':
            serializer = CreditoPersonasSerializer(query)
            createLog(logModel, serializer.data, logTransaccion)
            return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


# ACTUALIZAR
# 'methods' can be used to apply the same modification to multiple methods
@swagger_auto_schema(methods=['post'], request_body=CreditoPersonasSerializer)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def creditoPersonas_update(request, pk):
    """
    Este metod sirve para actualizar un creditopersona
    @type pk: recibe el id de la tabla credito persona
    @type request: recibe los campos de la tabla credito personas
    @rtype: DEvuelve el registro obtenido, caso contrario devuelve el error generado
    """
    request.POST._mutable = True
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'update/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'ESCRIBIR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            logModel['dataEnviada'] = str(request.data)
            query = CreditoPersonas.objects.filter(pk=ObjectId(pk), state=1).order_by('-created_at').first()
        except CreditoPersonas.DoesNotExist:
            errorNoExiste = {'error': 'No existe'}
            createLog(logModel, errorNoExiste, logExcepcion)
            return Response(status=status.HTTP_404_NOT_FOUND)
        if request.method == 'POST':
            now = timezone.localtime(timezone.now())
            request.data['updated_at'] = str(now)
            if 'created_at' in request.data:
                request.data.pop('created_at')

            if 'claveFirma' in request.data:
                if request.data['claveFirma'] != '':
                    usuario = query.empresaInfo
                    date = now.strftime("D:%Y%m%d%H%M%S+00'00'")
                    dct = {
                        "aligned": 0,
                        "sigflags": 3,
                        "sigflagsft": 132,
                        "sigpage": 0,
                        "sigbutton": True,
                        "sigfield": "Signature1",
                        "auto_sigfield": True,
                        "sigandcertify": True,
                        "signaturebox": (470, 840, 570, 640),
                        "signature": usuario['reprsentante'],
                        "signature_manual": [],
                        # "signature_img": "signature_test.png",
                        "contact": usuario['correo'],
                        "location": "Ubicación",
                        "signingdate": date,
                        "reason": "Firmar documentos habilitantes",
                        "password": request.data['claveFirma'],
                    }
            if 'solicitudCreditoFirmado' in request.data:
                print('entro')
                if request.data['solicitudCreditoFirmado'] is not None:
                    archivo_pdf_para_enviar_al_cliente = firmar(request, dct, 'solicitudCreditoFirmado')
                    request.data['solicitudCreditoFirmado'] = InMemoryUploadedFile(archivo_pdf_para_enviar_al_cliente,
                                                                                   'media',
                                                                                   'solicitudCreditoFirmado.pdf',
                                                                                   'application/pdf',
                                                                                   archivo_pdf_para_enviar_al_cliente.getbuffer().nbytes,
                                                                                   None
                                                                                   )
            if 'pagareFirmado' in request.data:
                if request.data['pagareFirmado'] is not None:
                    archivo_pdf_para_enviar_al_cliente = firmar(request, dct, 'pagareFirmado')
                    request.data['pagareFirmado'] = InMemoryUploadedFile(archivo_pdf_para_enviar_al_cliente,
                                                                         'media',
                                                                         'pagareFirmado.pdf',
                                                                         'application/pdf',
                                                                         archivo_pdf_para_enviar_al_cliente.getbuffer().nbytes,
                                                                         None
                                                                         )

            if 'contratosCuentaFirmado' in request.data:
                if request.data['contratosCuentaFirmado'] is not None:
                    archivo_pdf_para_enviar_al_cliente = firmar(request, dct, 'contratosCuentaFirmado')
                    request.data['contratosCuentaFirmado'] = InMemoryUploadedFile(archivo_pdf_para_enviar_al_cliente,
                                                                                  'media',
                                                                                  'contratosCuentaFirmado.pdf',
                                                                                  'application/pdf',
                                                                                  archivo_pdf_para_enviar_al_cliente.getbuffer().nbytes,
                                                                                  None
                                                                                  )
            if 'tablaAmortizacionFirmado' in request.data:
                if request.data['tablaAmortizacionFirmado'] is not None:
                    archivo_pdf_para_enviar_al_cliente = firmar(request, dct, 'tablaAmortizacionFirmado')
                    request.data['tablaAmortizacionFirmado'] = InMemoryUploadedFile(archivo_pdf_para_enviar_al_cliente,
                                                                                    'media',
                                                                                    'tablaAmortizacionFirmado.pdf',
                                                                                    'application/pdf',
                                                                                    archivo_pdf_para_enviar_al_cliente.getbuffer().nbytes,
                                                                                    None
                                                                                    )

            serializer = CreditoPersonasSerializer(query, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                createLog(logModel, serializer.data, logTransaccion)
                usuario = serializer.data['user']
                email = usuario['email'] if usuario else serializer.data['email']
                if email == '' or email is None:
                    email = serializer.data['empresaInfo']['correo']
                if serializer.data['estado'] == 'Negado':
                    if serializer.data['alcance'] != 'LOCAL':
                        # Publicar en la cola
                        publish(serializer.data)
                        if 'Pymes' in serializer.data['tipoCredito']:
                            enviarCorreoNegado(serializer.data['montoLiquidar'], email)
                        else:
                            enviarCorreoNegado(serializer.data['montoLiquidar'], email)
                if serializer.data['estado'] == 'Por Completar':
                    if serializer.data['alcance'] != 'LOCAL':
                        # Publicar en la cola
                        publish(serializer.data)
                        if 'Pymes' in serializer.data['tipoCredito']:
                            enviarCorreoPorcompletarLineaCredito(serializer.data['montoLiquidar'], email)
                        else:
                            enviarCorreoPorcompletar(serializer.data['montoLiquidar'], email)
                if serializer.data['estado'] == 'Aprobado' and 'motivoNegarLinea':
                    if serializer.data['montoLiquidar']:
                        if 'Pymes' in serializer.data['tipoCredito']:
                            if serializer.data['nombresCompleto']:
                                nombresCompleto = serializer.data['nombresCompleto']
                            else:
                                nombresCompleto = serializer.data['empresaInfo']['reprsentante']
                            enviarCorreoAprobado(serializer.data['montoAprobado'], email, nombresCompleto)
                        else:
                            enviarCorreoAprobadoCreditoConsumo(serializer.data['montoAprobado'], email)
                if 'tipoCredito' in request.data:
                    if request.data['tipoCredito'] == '':
                        credito = serializer.data
                        credito['tipoCredito'] = credito['canal']
                        if serializer.data['alcance'] != 'LOCAL':
                            # Publicar en la cola
                            publish(credito)
                        else:
                            query.tipoCredito = credito['canal']
                            query.save()
                if "estado" in request.data:
                    if request.data["estado"] == 'Enviado':
                        query.enviado = 1
                        query.save()
                if "motivoNegarLinea" in request.data and request.data['motivoNegarLinea'] != '':
                    enviarCorreoNegarLineaCredito(email, request.data['motivoNegarLinea'])
                if "activarMenu" in request.data and request.data['activarMenu']:
                    query.activarMenu = 1
                    query.save()

                return Response(serializer.data)
            createLog(logModel, serializer.errors, logExcepcion)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


# ELIMINAR
@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def creditoPersonas_delete(request, pk):
    """
    Este metod sirve para eliminar un creditopersona
    @type pk: recibe el id de la tabla credito persona
    @type request: no recibe nada
    @rtype: DEvuelve el registro eliminado, caso contrario devuelve el error generado
    """
    nowDate = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'delete/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'BORRAR',
        'fechaInicio': str(nowDate),
        'dataEnviada': '{}',
        'fechaFin': str(nowDate),
        'dataRecibida': '{}'
    }
    try:
        try:
            query = CreditoPersonas.objects.filter(user_id=pk, state=1).first()
        except CreditoPersonas.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'DELETE':
            serializer = CreditoPersonasSerializer(query, data={'state': '0', 'updated_at': str(nowDate)}, partial=True)
            if serializer.is_valid():
                serializer.save()
                createLog(logModel, serializer.data, logTransaccion)
                return Response(serializer.data, status=status.HTTP_200_OK)
            createLog(logModel, serializer.errors, logExcepcion)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


# ENCONTRAR CODIGO CREDITO PREAPROBADO
@api_view(['POST'])
def creditoPersonas_creditoPreaprobado_codigo(request):
    """
    Este metodo sirve para consultar un credito por el codigo
    @type request: recibe el codigo, cedula, rucEmpresa
    @rtype: devuelve el registro, caso contrario devuelve el error generado
    """
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'creditoPreaprobado/codigo',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)
            # Filtros
            filters = {"state": "1"}

            if "codigo" in request.data:
                if request.data["codigo"] != '':
                    filters['codigoPreaprobado'] = request.data["codigo"]

            if "cedula" in request.data:
                if request.data["cedula"] != '':
                    filters['numeroIdentificacion'] = request.data["cedula"]

            if "rucEmpresa" in request.data:
                if request.data["rucEmpresa"] != '':
                    filters['rucEmpresa'] = request.data["rucEmpresa"]

            # Serializar los datos
            try:
                query = CreditoPersonas.objects.get(**filters)
            except CreditoPersonas.DoesNotExist:
                err = {"error": "No existe"}
                createLog(logModel, err, logExcepcion)
                return Response(err, status=status.HTTP_404_NOT_FOUND)

            # response = {'monto': query.monto, 'nombreCompleto': query.nombres + ' ' + query.apellidos,
            #             'tipoPersona': query.tipoPersona, 'estadoCivil': query.estadoCivil}
            # query.state = 0
            # query.save()

            # envio de datos
            serializer = CreditoPersonasSerializer(query)
            createLog(logModel, serializer.data, logTransaccion)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def creditoPersonas_list(request):
    """
    Este metodo sirve para listar los
    @type request: recibe page, page_size, empresaComercial_id, empresaIfis_id, estado, tipoCredito, user_id, canal, cargarOrigen, enviado
    @rtype: DEvuelve una lista, caso contrario devuelve el error generado
    """
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'list/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)
            # paginacion
            page_size = int(request.data['page_size'])
            page = int(request.data['page'])
            offset = page_size * page
            limit = offset + page_size
            # Filtros
            filters = {"state": "1"}

            if "empresaComercial_id" in request.data:
                if request.data["empresaComercial_id"] != '':
                    filters['empresaComercial_id'] = ObjectId(request.data["empresaComercial_id"])

            if "empresaIfis_id" in request.data:
                if request.data["empresaIfis_id"] != '':
                    filters['empresaIfis_id'] = ObjectId(request.data["empresaIfis_id"])

            if "estado" in request.data:
                if request.data["estado"] != '':
                    filters['estado__icontains'] = request.data["estado"]

            if "tipoCredito" in request.data:
                if request.data["tipoCredito"] != '':
                    filters['tipoCredito'] = str(request.data["tipoCredito"])

            if "user_id" in request.data:
                if request.data["user_id"] != '':
                    filters['user_id'] = str(request.data["user_id"])

            if "canal" in request.data:
                if request.data["canal"] != '':
                    filters['canal'] = str(request.data["canal"])

            if "numeroIdentificacion" in request.data:
                if request.data["numeroIdentificacion"] != '':
                    filters['numeroIdentificacion'] = str(request.data["numeroIdentificacion"])

            if "cargarOrigen" in request.data:
                if request.data["cargarOrigen"] != '':
                    filters['cargarOrigen__icontains'] = str(request.data["cargarOrigen"])

            if "alcance" in request.data:
                if request.data["alcance"] != '':
                    filters['alcance__icontains'] = request.data["alcance"]

            if "enviado" in request.data:
                filters['enviado'] = request.data["enviado"]

            # Serializar los datos
            query = CreditoPersonas.objects.filter(**filters).order_by('-created_at')
            serializer = CreditoPersonasSerializer(query[offset:limit], many=True)
            new_serializer_data = {'cont': query.count(),
                                   'info': serializer.data}
            # envio de datos
            return Response(new_serializer_data, status=status.HTTP_200_OK)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)


# METODO SUBIR ARCHIVOS EXCEL
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def uploadEXCEL_creditosPreaprobados(request):
    """
    ESte metodo sirve para cargar el archivo de credito preaprobados empleados
    @type pk: el id de la tabla creditopreaprobado empleados
    @type request: no recibe nada
    @rtype: Devuelve los registros correctos, incorrectos, caso contrario devuelve el error generado
    """
    contValidos = 0
    contInvalidos = 0
    contTotal = 0
    errores = []
    try:
        if request.method == 'POST':
            first = True  # si tiene encabezado
            uploaded_file = request.FILES['documento']
            # you may put validations here to check extension or file size
            wb = openpyxl.load_workbook(uploaded_file)
            # getting a particular sheet by name out of many sheets
            worksheet = wb["Clientes"]
            lines = list()
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            lines.append(row_data)

        for dato in lines:
            contTotal += 1
            if first:
                first = False
                continue
            else:
                if len(dato) == 7:
                    resultadoInsertar = insertarDato_creditoPreaprobado(dato, request.data['empresa_financiera'])
                    if resultadoInsertar != 'Dato insertado correctamente':
                        contInvalidos += 1
                        errores.append({"error": "Error en la línea " + str(contTotal) + ": " + str(resultadoInsertar)})
                    else:
                        contValidos += 1
                else:
                    contInvalidos += 1
                    errores.append({"error": "Error en la línea " + str(
                        contTotal) + ": la fila tiene un tamaño incorrecto (" + str(len(dato)) + ")"})

        result = {"mensaje": "La Importación se Realizo Correctamente",
                  "correctos": contValidos,
                  "incorrectos": contInvalidos,
                  "errores": errores
                  }
        return Response(result, status=status.HTTP_201_CREATED)

    except Exception as e:
        err = {"error": 'Error verifique el archivo, un error ha ocurrido: {}'.format(e)}
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


# INSERTAR DATOS EN LA BASE INDIVIDUAL
def insertarDato_creditoPreaprobado(dato, empresa_financiera):
    """
    ESte metodo sirve para enviar el correo
    @param dato: recibe la fila del excel
    @param empresa_financiera: recibe la empresa
    @param empresa_comercial: recibe la empresa
    @rtype: No devuelve nada
    """
    try:
        timezone_now = timezone.localtime(timezone.now())
        data = {}
        data['vigencia'] = dato[0].replace('"', "")[0:10] if dato[0] != "NULL" else None
        data['concepto'] = dato[1].replace('"', "") if dato[1] != "NULL" else None
        data['monto'] = dato[2].replace('"', "") if dato[2] != "NULL" else None
        data['plazo'] = dato[3].replace('"', "") if dato[3] != "NULL" else None
        data['interes'] = dato[4].replace('"', "") if dato[4] != "NULL" else None
        data['estado'] = 'PreAprobado'
        data['tipoCredito'] = 'PreAprobado'
        data['canal'] = 'PreAprobado'
        persona = Personas.objects.filter(identificacion=dato[5], state=1).first()
        data['user_id'] = persona.user_id
        data['numeroIdentificacion'] = dato[5]
        data['nombres'] = persona.nombres
        data['apellidos'] = persona.apellidos
        data['nombresCompleto'] = persona.nombres + ' ' + persona.apellidos
        data['empresaIfis_id'] = empresa_financiera
        data['empresasAplican'] = dato[6]
        data['created_at'] = str(timezone_now)
        # inserto el dato con los campos requeridos
        CreditoPersonas.objects.create(**data)
        return 'Dato insertado correctamente'
    except Exception as e:
        return str(e)


# METODO SUBIR ARCHIVOS EXCEL
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def uploadEXCEL_creditosPreaprobados_empleados(request):
    """
    ESte metodo sirve para cargar el archivo de credito preaprobados empleados
    @type pk: el id de la tabla creditopreaprobado empleados
    @type request: no recibe nada
    @rtype: Devuelve los registros correctos, incorrectos, caso contrario devuelve el error generado
    """
    contValidos = 0
    contInvalidos = 0
    contTotal = 0
    errores = []
    try:
        if request.method == 'POST':
            first = True  # si tiene encabezado
            uploaded_file = request.FILES['documento']
            # you may put validations here to check extension or file size
            wb = openpyxl.load_workbook(uploaded_file)
            # getting a particular sheet by name out of many sheets
            worksheet = wb["Clientes"]
            lines = list()
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            lines.append(row_data)

        for dato in lines:
            contTotal += 1
            if first:
                first = False
                continue
            else:
                if len(dato) == 11:
                    resultadoInsertar = insertarDato_creditoPreaprobado_empleado(dato,
                                                                                 request.data['empresa_financiera'])
                    if resultadoInsertar != 'Dato insertado correctamente':
                        contInvalidos += 1
                        errores.append({"error": "Error en la línea " + str(contTotal) + ": " + str(resultadoInsertar)})
                    else:
                        contValidos += 1
                else:
                    contInvalidos += 1
                    errores.append({"error": "Error en la línea " + str(
                        contTotal) + ": la fila tiene un tamaño incorrecto (" + str(len(dato)) + ")"})

        result = {"mensaje": "La Importación se Realizo Correctamente",
                  "correctos": contValidos,
                  "incorrectos": contInvalidos,
                  "errores": errores
                  }
        return Response(result, status=status.HTTP_201_CREATED)

    except Exception as e:
        err = {"error": 'Error verifique el archivo, un error ha ocurrido: {}'.format(e)}
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


# INSERTAR DATOS EN LA BASE INDIVIDUAL
def insertarDato_creditoPreaprobado_empleado(dato, empresa_financiera):
    """
    ESte metodo sirve para enviar el correo
    @param dato: recibe la fila del excel
    @param empresa_financiera: recibe la empresa
    @param empresa_comercial: recibe la empresa
    @rtype: No devuelve nada
    """
    try:
        timezone_now = timezone.localtime(timezone.now())
        data = {}
        data['vigencia'] = dato[0].replace('"', "")[0:10] if dato[0] != "NULL" else None
        data['concepto'] = dato[1].replace('"', "") if dato[1] != "NULL" else None
        data['monto'] = dato[2].replace('"', "") if dato[2] != "NULL" else None
        data['plazo'] = dato[3].replace('"', "") if dato[3] != "NULL" else None
        data['interes'] = dato[4].replace('"', "") if dato[4] != "NULL" else None
        data['estado'] = 'PreAprobado'
        data['tipoCredito'] = ''
        data['canal'] = 'Empleado'
        persona = Personas.objects.filter(identificacion=dato[5], state=1).first()
        data['user_id'] = persona.user_id
        data['numeroIdentificacion'] = dato[5]
        data['nombres'] = persona.nombres
        data['apellidos'] = persona.apellidos
        data['nombresCompleto'] = persona.nombres + ' ' + persona.apellidos
        data['empresaIfis_id'] = empresa_financiera
        data['empresasAplican'] = dato[10]
        data['created_at'] = str(timezone_now)
        # inserto el dato con los campos requeridos
        CreditoPersonas.objects.create(**data)
        return 'Dato insertado correctamente'
    except Exception as e:
        return str(e)


# ENCONTRAR UNO
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def creditoPersonas_listOne_persona(request, pk):
    """
    Este metod sirve para obtener un creditopersona
    @type pk: recibe el id de la tabla credito persona
    @type request: no recibe
    @rtype: DEvuelve el registro obtenido, caso contrario devuelve el error generado
    """
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'listOne/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            query = CreditoPersonas.objects.filter(pk=ObjectId(pk), state=1).first()
        except CreditoPersonas.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'GET':
            serializer = CreditoPersonasPersonaSerializer(query)
            createLog(logModel, serializer.data, logTransaccion)
            return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def creditoPersonas_listOne_usuario(request, pk):
    """
    Este metod sirve para obtener un creditopersona
    @type pk: recibe el id de la tabla credito persona
    @type request: no recibe
    @rtype: DEvuelve el registro obtenido, caso contrario devuelve el error generado
    """
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'listOne/usuario/' + pk,
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            filters = {'state': 1}
            filters['user_id'] = pk
            filters['estado__icontains'] = request.data['estado']
            query = CreditoPersonas.objects.filter(**filters).first()
        except CreditoPersonas.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'POST':
            serializer = CreditoPersonasSerializer(query)
            createLog(logModel, serializer.data, logTransaccion)
            return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
# @permission_classes([IsAuthenticated])
def creditoPersonas_lecturaArchivos(request, pk):
    """
    Este metodo sirve para obtener los datos de un archivo
    @type pk: recibe el id del credito
    @type request: no recibe nada
    @rtype: devuelve los datos del documento, caso contrario devuelve el error generado
    """
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'lecturaArchivos/' + pk,
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            query = CreditoPersonas.objects.filter(pk=ObjectId(pk), state=1).first()
        except CreditoPersonas.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'GET':
            print(query.identificacion.name)
            dato1 = None if query.identificacion.name is None else obtenerDatosArchivos(str(query.identificacion.name))
            dato2 = None if query.ruc.name is None else obtenerDatosArchivos(str(query.ruc.name))
            # serializer = CreditoPersonasPersonaSerializer(query)
            # createLog(logModel, serializer.data, logTransaccion)
            return Response({'cedula': dato1, 'ruc': dato2}, status=status.HTTP_200_OK)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


def obtenerDatosArchivos(nombreArchivo):
    """
    Este metodo sirve para leer un documento pdf
    @type nombreArchivo: recibe el nombre del archivo a leer
    @rtype: DEvuelve los datos del archivo, caso contrario devuelve el error generado
    """
    # environ init
    env = environ.Env()
    environ.Env.read_env()  # LEE ARCHIVO .ENV
    # Function invokes
    jobId = InvokeTextDetectJob(env.str('AWS_STORAGE_BUCKET_NAME'), nombreArchivo)
    print("Started job with id: {}".format(jobId))
    respuesta = {}
    if (CheckJobComplete(jobId)):
        response = JobResults(jobId)
        for resultPage in response:
            for item in resultPage["Blocks"]:
                if item['BlockType'] == 'LINE':
                    if re.match("\d{10}001", item['Text']):
                        respuesta['ruc'] = item['Text']

                    elif re.match("No. \d{9}-[0-9]", item['Text']):
                        respuesta['identificacion'] = item['Text'][4:]

                    elif re.match("^\d{4}\-(0?[1-9]|1[012])\-(0?[1-9]|[12][0-9]|3[01])$", item['Text']):
                        respuesta['fechaExpiracion'] = item['Text']

                    elif re.match("[aA-Zz]\d{4}[aA-Zz]\d{4}", item['Text']):
                        respuesta['codigoDactilar'] = item['Text']

    print("-------------------Imprimir-----------------")
    return respuesta


import time


## Textract APIs used - "start_document_text_detection", "get_document_text_detection"
def InvokeTextDetectJob(bucket, nombreArchivo):
    response = None
    textarctmodule = boto3.client(
        'textract',
        aws_access_key_id=config.AWS_ACCESS_KEY_ID_TEXTRACT,
        aws_secret_access_key=config.AWS_SECRET_ACCESS_KEY_TEXTRACT,
        region_name='us-east-1'
    )
    response = textarctmodule.start_document_text_detection(
        DocumentLocation={
            'S3Object': {
                'Bucket': bucket,
                # 'Name': nombreArchivo
                'Name': 'CORP/documentosCreditosPersonas/62d97613bceeaa781e803920_1658498310065_comprobante_1.pdf'
            }
        }
    )
    return response["JobId"]


def CheckJobComplete(jobId):
    time.sleep(5)
    client = boto3.client(
        'textract',
        aws_access_key_id=config.AWS_ACCESS_KEY_ID_TEXTRACT,
        aws_secret_access_key=config.AWS_SECRET_ACCESS_KEY_TEXTRACT,
        region_name='us-east-1'
    )
    response = client.get_document_text_detection(JobId=jobId)
    status = response["JobStatus"]
    print("Job status: {}".format(status))
    while (status == "IN_PROGRESS"):
        time.sleep(5)
        response = client.get_document_text_detection(JobId=jobId)
        status = response["JobStatus"]
        print("Job status: {}".format(status))
    return status


def JobResults(jobId):
    pages = []
    client = boto3.client(
        'textract',
        aws_access_key_id=config.AWS_ACCESS_KEY_ID_TEXTRACT,
        aws_secret_access_key=config.AWS_SECRET_ACCESS_KEY_TEXTRACT,
        region_name='us-east-1'
    )
    response = client.get_document_text_detection(JobId=jobId)

    pages.append(response)
    print("Resultset page recieved: {}".format(len(pages)))
    nextToken = None
    if ('NextToken' in response):
        nextToken = response['NextToken']
        while (nextToken):
            response = client.get_document_text_detection(JobId=jobId, NextToken=nextToken)
            pages.append(response)
            print("Resultset page recieved: {}".format(len(pages)))
            nextToken = None
            if ('NextToken' in response):
                nextToken = response['NextToken']
    return pages


@api_view(['GET'])
def prueba(request):
    """
    Este metodo sirve para actualizar la cola aws
    @type request: no recibe nada
    @rtype: No devuelve nada
    """
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'listOne/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        get_queue_url()
        msg = {"msg": "Se actualizo la cola"}
        return Response(msg, status=status.HTTP_202_ACCEPTED)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


def enviarCodigoCorreoMicroCredito(codigo, email, nombreSolicitante):
    """
    Este metodo sirve para enviar correo
    @type email: Recibe el email
    @type codigo: recibe el codigo
    @type nombreSolicitante: recibe el nombre del solicitante
    """
    subject, from_email, to = 'Código de seguridad para Consulta de Crédito', "08d77fe1da-d09822@inbox.mailtrap.io", \
                              email
    txt_content = f"""
        Se ha generado su código para consultar su crédito
                            
       Estimad@ {nombreSolicitante}
        
        El código para realizar la consulta de su crédito es: {codigo}
        
        Por su seguridad, comparta esta información únicamente con el vendedor del local comercial en el que desea 
        realizar su compra. Una vez ingresado, bloquearemos otras consultas para que su compra sea realizada de manera segura.
        
        Atentamente,
        
        CrediCompra-Big Puntos.
    """
    html_content = f"""
        <html>
            <body>
                <h1>
                Se ha generado su código para consultar su crédito
                </h1>
                <br>
                <p>
                Estimad@ {nombreSolicitante}
                </p>
                <br>
                <p>
                El código para realizar la consulta de su crédito es: {codigo}
                </p>
                <br>
                <p>
                Por su seguridad, comparta esta información únicamente con el vendedor del local comercial en el que
                 desea realizar su compra. Una vez ingresado, bloquearemos otras consultas para que su compra sea realizada de manera segura.
                </p>
                Atentamente,
                <br>
                CrediCompra-Big Puntos.
                <br>
            </body>
        </html>
    """
    sendEmail(subject, txt_content, from_email, to, html_content)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def creditoPersonas_codigo_creditoAprobado(request):
    """
    Este metodo sirve para buscar el credito aprobado
    @type request: recibe numeroIdentificacion
    @rtype: deveulve el credito aprobado, caso contrario devuelve el error generado
    """
    print('metodo')
    nowDate = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'generar/codigo/creditoAprobado/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'BORRAR',
        'fechaInicio': str(nowDate),
        'dataEnviada': '{}',
        'fechaFin': str(nowDate),
        'dataRecibida': '{}'
    }
    try:
        try:
            query = CreditoPersonas.objects.filter(numeroIdentificacion=request.data['numeroIdentificacion'], state=1,
                                                   estado='Aprobado').order_by('-created_at').first()
        except CreditoPersonas.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'POST':
            # Genera el codigo
            longitud_codigo = Catalogo.objects.filter(tipo='CONFIG_TWILIO', nombre='LONGITUD_CODIGO',
                                                      state=1).first().valor
            codigo = (''.join(random.choice(string.digits) for _ in range(int(longitud_codigo))))
            nombreSolicitante = query.user['nombres'] + ' ' + query.user['apellidos']
            enviarCodigoCorreoMicroCredito(codigo, query.user['email'], nombreSolicitante)
            serializer = CodigoCreditoSerializer(data={'credito_id': str(query._id), 'codigo': codigo,
                                                       'numeroIdentificacion': request.data['numeroIdentificacion']})
            if serializer.is_valid():
                serializer.save()
                createLog(logModel, serializer.data, logTransaccion)
                return Response(serializer.data, status=status.HTTP_200_OK)
            createLog(logModel, serializer.errors, logExcepcion)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def creditoPersonas_validar_codigo_creditoAprobado(request):
    """
    Este metodo sirve para validar el credito aprobado
    @type request: recibe numeroIdentificacion, codigo
    @rtype: deveulve el credito aprobado, caso contrario devuelve el error generado
    """
    nowDate = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'validar/codigo/creditoAprobado',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'BORRAR',
        'fechaInicio': str(nowDate),
        'dataEnviada': '{}',
        'fechaFin': str(nowDate),
        'dataRecibida': '{}'
    }
    try:
        try:
            query = CodigoCredito.objects.filter(numeroIdentificacion=request.data['numeroIdentificacion'], state=1,
                                                 codigo=request.data['codigo']).first()
        except CodigoCredito.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'POST':
            tiempo = Catalogo.objects.filter(tipo='CONFIG_DURACION_CREDITO_APROBADO',
                                             nombre='DURACION_CODIGO_CREDITO_APROBADO', state=1).first().valor
            duracion = query.created_at + relativedelta(minutes=int(tiempo))
            if duracion > timezone.now():
                credito = CreditoPersonas.objects.get(_id=ObjectId(query.credito_id))
                serializer = CreditoPersonasSerializer(credito)
                createLog(logModel, serializer.data, logTransaccion)
                return Response(serializer.data, status=status.HTTP_200_OK)
            else:
                serializer = {'mensaje': 'No existe'}
                createLog(logModel, serializer, logExcepcion)
                return Response(serializer, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


def enviarCorreoSolicitud(email):
    """
    Este metodo sirve para enviar el corrreo
    @type email: Recibe el email
    @rtype: no devuelve nada
    """
    subject, from_email, to = 'Solicitud de Crédito Recibida – Crédito Pagos', "08d77fe1da-d09822@inbox.mailtrap.io", \
                              email
    txt_content = f"""
                        Global RedPyme - Crédito Pagos ha recibido su solicitud, estaremos en contacto con usted a la brevedad posible.
                            Crédito Pagos es la mejor opción para el crecimiento de su negocio
                        Atentamente,
                        Global RedPyme – Crédito Pagos
    """
    html_content = f"""
                <html>
                    <body>
                        <p>Global RedPyme - Crédito Pagos ha recibido su solicitud, estaremos en contacto con usted a la brevedad posible.</p>
                        <br>
                        <br>
                        <p><b>Crédito Pagos es la mejor opción para el crecimiento de su negocio</b></p>
                        <br>
                        Atentamente,
                        <br>
                        Global RedPyme – Crédito Pagos
                        <br>
                    </body>
                </html>
                """
    sendEmail(subject, txt_content, from_email, to, html_content)


def firmar(request, dct, nombreArchivo):
    """
    Este metodo sirve para firmar un documento
    @type nombreArchivo: recibe el nombre
    @type dct: recibe los parametros para la firma
    @type request: recibe los datos enviados para la tabla creditos personas
    @rtype: devuelve el archivo firmado
    """
    certificado = request.data['certificado']
    # environ init
    env = environ.Env()
    environ.Env.read_env()  # LEE ARCHIVO .ENV
    client_s3 = boto3.client(
        's3',
        aws_access_key_id=env.str('AWS_ACCESS_KEY_ID'),
        aws_secret_access_key=env.str('AWS_SECRET_ACCESS_KEY')
    )
    with tempfile.TemporaryDirectory() as d:
        ruta = d + 'SOLICITUD_REMATRICULA_DE_.pdf'
        s3 = boto3.resource('s3')
        archivo = s3.meta.client.download_file(env.str('AWS_STORAGE_BUCKET_NAME'),
                                               urllib.parse.unquote(str(request.data[nombreArchivo])), ruta)
    date = datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S")
    datosFirmante = f"""FIRMADO POR:\n {dct['signature']} \n FECHA:\n {date}"""
    generarQR(datosFirmante)
    output_file = "example-with-barcode.pdf"
    agregarQRDatosFirmante(datosFirmante, output_file, ruta)

    contrasenia = request.data['claveFirma']
    p12 = pkcs12.load_key_and_certificates(
        certificado.read(), contrasenia.encode("ascii"), backends.default_backend()
    )
    datau = open(output_file, "rb").read()
    dct.pop('signature')
    datas = cms.sign(datau, dct, p12[0], p12[1], p12[2], "sha256")
    archivo_pdf_para_enviar_al_cliente = io.BytesIO()
    archivo_pdf_para_enviar_al_cliente.write(datau)
    archivo_pdf_para_enviar_al_cliente.write(datas)
    archivo_pdf_para_enviar_al_cliente.seek(0)
    return archivo_pdf_para_enviar_al_cliente


def enviarCorreoNegado(montoAprobado, email):
    """
    Este metodo sirve para enviar el correo de credito negado
    @type email: recibe el email
    @type montoAprobado: recibe el monto aprobado
    @rtype: no devuelve nada
    """
    subject, from_email, to = 'Solicitud de Línea de Crédito NEGADA', "08d77fe1da-d09822@inbox.mailtrap.io", \
                              email
    txt_content = f"""
        <h1>¡LO SENTIMOS!</h1>
                            
        Su Solicitud de Línea de Crédito para realizar pagos a sus proveedores otorgado por una Cooperativa de 
        Ahorro y Crédito regulada ha sido NEGADA.
        
        Contáctese con su asesor a través de nuestro link de WhatsApp: https://wa.link/e8b3sa
        
        Atentamente,
        
        Equipo Global Redpyme – Crédito Pagos
    """
    html_content = f"""
                <html>
                    <body>
                        <h1>¡LO SENTIMOS!</h1>
                        <br>
                        <p>
                        Su Solicitud de Línea de Crédito para realizar pagos a sus proveedores otorgado por una 
                        Cooperativa de Ahorro y Crédito regulada ha sido <b>NEGADA</b>.
                        </p>
                        <p>
                        Contáctese con su asesor a través de nuestro link de WhatsApp: <a href='https://wa.link/e8b3sa'>LINK</a>
                        </p>
                        <br>
                        Atentamente,
                        <br>
                        Equipo Global Redpyme – Crédito Pagos
                        <br>
                    </body>
                </html>
                """
    sendEmail(subject, txt_content, from_email, to, html_content)


def enviarCorreoNegado(montoAprobado, email):
    """
    Este metodo sirve para enviar el correo de credito negado
    @type email: recibe el email
    @type montoAprobado: recibe el monto aprobado
    @rtype: no devuelve nada
    """
    subject, from_email, to = 'Solicitud de Línea de Crédito NEGADA', "08d77fe1da-d09822@inbox.mailtrap.io", \
                              email
    txt_content = f"""
        <h1>¡LO SENTIMOS!</h1>

        Su solicitud de crédito para realizar compras en las mejores Casas Comerciales con un crédito otorgado 
        por una Cooperativa de Ahorro y Crédito regulada ha sido NEGADA.

        Contáctese con su asesor a través de nuestro link de WhatsApp: https://wa.link/e8b3sa

        Atentamente,

        Equipo Global Redpyme – Crédito Pagos
    """
    html_content = f"""
                <html>
                    <body>
                        <h1>¡LO SENTIMOS!</h1>
                        <br>
                        <p>
                        Su solicitud de crédito para realizar compras en las mejores Casas Comerciales con un crédito
                         otorgado por una Cooperativa de Ahorro y Crédito regulada ha sido <b>NEGADA</b>.
                        </p>
                        <p>
                        Contáctese con su asesor a través de nuestro link de WhatsApp: <a href='https://wa.link/e8b3sa'>LINK</a>
                        </p>
                        <br>
                        Atentamente,
                        <br>
                        Equipo Global Redpyme – Crédito Pagos
                        <br>
                    </body>
                </html>
                """
    sendEmail(subject, txt_content, from_email, to, html_content)


def enviarCorreoPorcompletar(montoAprobado, email):
    """
    Este metodo sirve para enviar el correo de credito negado
    @type email: recibe el email
    @type montoAprobado: recibe el monto aprobado
    @rtype: no devuelve nada
    """
    subject, from_email, to = 'Solicitud de Crédito DEVUELTA para Completar información', "08d77fe1da-d09822@inbox.mailtrap.io", \
                              email
    txt_content = f"""
        ¡LO SENTIMOS!
                                                    
        Su Solicitud de Crédito para realizar compras en los mejores Locales Comerciales del país con un Crédito 
        otorgado por una Cooperativa de Ahorro y Crédito regulada, ha sido DEVUELTA para completar información.
                                
        Contáctese con su asesor a través de nuestro link de WhatsApp: https://wa.link/szsyad
        
        Atentamente,
        
        CrediCompra – Big Puntos
    """
    html_content = f"""
                <html>
                    <body>
                        <h1>¡LO SENTIMOS!</h1>
                        <br>
                        <p>
                        Su Solicitud de Crédito para realizar compras en los mejores Locales Comerciales del país con 
                        un Crédito otorgado por una Cooperativa de Ahorro y Crédito regulada, 
                        ha sido DEVUELTA para completar información.
                        </p>
                        <br>
                        <p>Contáctese con su asesor a través de nuestro link de WhatsApp: <a href='https://wa.link/szsyad'>LINK</a></p>
                        <br>
                        Atentamente,
                        <br>
                        CrediCompra – Big Puntos
                        <br>
                    </body>
                </html>
                """
    sendEmail(subject, txt_content, from_email, to, html_content)


def enviarCorreoPorcompletarLineaCredito(montoAprobado, email):
    """
    Este metodo sirve para enviar el correo de credito negado
    @type email: recibe el email
    @type montoAprobado: recibe el monto aprobado
    @rtype: no devuelve nada
    """
    subject, from_email, to = 'Solicitud de Línea de Crédito DEVUELTA para completar información', "08d77fe1da-d09822@inbox.mailtrap.io", \
                              email
    txt_content = f"""
        ¡LO SENTIMOS!

        Su Solicitud de Línea de Crédito para realizar pagos a sus proveedores y/o empleados con una Línea de Crédito
        otorgada por una Cooperativa de Ahorro y Crédito regulada ha sido DEVUELTA PARA COMPLETAR INFORMACIÓN.

        Contáctese con su asesor a través de nuestro link de WhatsApp https://wa.link/s3vift

        Atentamente,

        Equipo Global Redpyme – Crédito Pagos
    """
    html_content = f"""
        <html>
            <body>
                <h1>¡LO SENTIMOS!</h1>
                <br>
                <p>
                Su Solicitud de Línea de Crédito para realizar pagos a sus proveedores y/o empleados con una Línea
                de Crédito otorgada por una Cooperativa de Ahorro y Crédito regulada ha sido DEVUELTA PARA COMPLETAR INFORMACIÓN.
                </p>
                <br>
                <p>Contáctese con su asesor a través de nuestro link de WhatsApp <a href='https://wa.link/s3vift'>LINK</a></p>
                <br>
                Atentamente,
                <br>
                Equipo Global Redpyme – Crédito Pagos
                <br>
            </body>
        </html>
    """
    sendEmail(subject, txt_content, from_email, to, html_content)


def enviarCorreoAprobado(montoAprobado, email, nombreCompleto):
    """
    Este metodo sirve para enviar el correo de credito negado
    @type email: recibe el email
    @type montoAprobado: recibe el monto aprobado
    @rtype: no devuelve nada
    """
    subject, from_email, to = 'Su Solicitud de Línea de Crédito ha sido APROBADA', "08d77fe1da-d09822@inbox.mailtrap.io", email
    txt_content = f"""
    LÍNEA DE CRÉDITO PARA PAGO A PROVEEDORES Y/O EMPLEADOS APROBADA
    
    Felicidades!
    Su Solicitud de Línea de Crédito para realizar pagos a sus proveedores y/o empleados ha sido APROBADA por un monto de {montoAprobado}.
    
    Para acceder a su Línea de Crédito para realizar pagos a proveedores y/o empleados, realice lo siguiente: 
    
    Ingrese a: {config.API_FRONT_END_CENTRAL}/grp/registro?email={email}&nombre={nombreCompleto}  y cargue su firma electrónica en 
    nuestra plataforma. Recuerde que al hacerlo, autoriza a la Plataforma y Cooperativa de Ahorro 
    y Crédito a realizar movimientos desde su cuenta con el único fin de completar el proceso del PAGO A SUS PROVEEDORES Y/O EMPLEADOS.
    
    Si requiere asistencia personalizada, contáctenos a través del siguiente <a href='https://wa.link/pcs3ll'>LINK</a>
    
    Atentamente,
    
    Global RedPyme – Crédito Pagos
    """
    html_content = f"""
                <html>
                    <body>
                        <h1>LÍNEA DE CRÉDITO PARA PAGO A PROVEEDORES Y/O EMPLEADOS APROBADA</h1>
                        <br>
                        <h2>Felicidades!</h2>
                        <p>
                        Su Solicitud de Línea de Crédito para realizar pagos a sus proveedores y/o empleados ha sido 
                        <b>APROBADA</b> por un monto de  {montoAprobado}.
                        </p>
                        <br>
                        <p>
                        <b>Para acceder a su Línea de Crédito para realizar pagos a proveedores y/o empleados, realice lo siguiente:</b>
                        </p>
                        <br>
                        <p>
                        Ingrese a: <a href="{config.API_FRONT_END_CENTRAL}/grp/registro?email={email}&nombre={nombreCompleto}">{config.API_FRONT_END_CENTRAL}/grp/registro?email={email}&nombre={nombreCompleto}</a> 
                         y cargue su firma electrónica en nuestra plataforma. Recuerde que al hacerlo, autoriza a la Plataforma y Cooperativa de Ahorro 
                        y Crédito a realizar movimientos desde su cuenta con el único fin de completar el proceso del PAGO A SUS PROVEEDORES Y/O EMPLEADOS.
                        </p>
                        <br>
                        <p>Si requiere asistencia personalizada, contáctenos a través del siguiente <a href='https://wa.link/pcs3ll'>LINK</a></p>
                        <br>
                        Atentamente,
                        <br>
                        <b>Global RedPyme – Crédito Pagos</b>
                        <br>
                    </body>
                </html>
                """
    sendEmail(subject, txt_content, from_email, to, html_content)
    print(email)


def enviarCorreoAprobadoCreditoConsumo(montoAprobado, email):
    """
    Este metodo sirve para enviar el correo de credito negado
    @type email: recibe el email
    @type montoAprobado: recibe el monto aprobado
    @rtype: no devuelve nada
    """
    subject, from_email, to = 'Su Solicitud de crédito de consumo ha sido APROBADA', "08d77fe1da-d09822@inbox.mailtrap.io", email
    txt_content = f"""
    CRÉDITO DE CONSUMO APROBADO

    Felicidades!
    Su Solicitud de Crédito para realizar compras en los mejores Locales Comerciales del país ha sido APROBADA por un monto de {montoAprobado}.

    Para acceder a su Crédito, realice los siguientes pasos:

    Ingrese a www.credicompra.com y revise el catálogo de nuestros Locales Comerciales afiliados.
    Acérquese al Local Comercial de su preferencia y solicite realizar la compra con su crédito Aprobado.
    Confirme sus datos
    Escoja sus productos y listo. Pague con su Crédito Aprobado

    Si requiere asistencia personalizada, contáctenos a través del siguiente <a href='https://wa.link/6m3c3k'>LINK</a>

    Atentamente,

    CrediCompra – Big Puntos
    """
    html_content = f"""
                <html>
                    <body>
                        <h1>CRÉDITO DE CONSUMO APROBADO</h1>
                        <br>
                        <h2>Felicidades!</h2>
                        <p>
                        Su Solicitud de Crédito para realizar compras en los mejores Locales Comerciales del país ha 
                        sido APROBADA por un monto de {montoAprobado}.
                        </p>
                        <br>
                        <p>
                        <b>Para acceder a su Crédito, realice los siguientes pasos:</b>
                        </p>
                        <br>
                        <ol>
                             <li>Ingrese a <a href='https://credicompra.com/'>www.credicompra.com</a> y revise el catálogo de nuestros Locales Comerciales afiliados.</li>
                             <li>Acérquese al Local Comercial de su preferencia y solicite realizar la compra con su crédito Aprobado.</li>
                             <li>Confirme sus datos</li>
                             <li>Escoja sus productos y listo. Pague con su Crédito Aprobado</li>
                        </ol>
                        <br>
                        <p>Si requiere asistencia personalizada, contáctenos a través del siguiente <a href='https://wa.link/ox0xha'>LINK</a></p>
                        <br>
                        Atentamente,
                        <br>
                        CrediCompra – Big Puntos
                        <br>
                    </body>
                </html>
                """
    sendEmail(subject, txt_content, from_email, to, html_content)
    print(email)


def generarQR(datos):
    """
    Este metodo sirve para generar el codigo qr
    @type datos: recibe los datos que se agregaran al qr
    @rtype: no devuelve nada
    """
    img = qrcode.make(datos)
    f = open("output.png", "wb")
    img.save(f)
    f.close()


def agregarQRDatosFirmante(datosFirmante, output_file, ruta):
    """
    Este metodo sirve para agregar el qr a un documento pdf
    @type ruta: recibe la ruta de donde esta el qr
    @type output_file: recibe la ruta donde se guarda el documento
    @type datosFirmante: recibe los datos del cliente
    @rtype: no devuelve nada
    """
    # Define the position and size of the image rectangle
    image_rectangle = fitz.Rect(0, 0, 250, 220)  # Adjust the coordinates and size as needed

    # Retrieve the first page of the PDF
    file_handle = fitz.open(ruta)
    first_page = file_handle[0]

    # Open and flip the image vertically
    image_path = 'output.png'
    image = Image.open(image_path).transpose(Image.FLIP_TOP_BOTTOM)
    image.save('flipped_image.png')

    # Insert the flipped image into the PDF
    img = open('output.png', "rb").read()  # an image file
    img_xref = 0
    first_page.insert_image(image_rectangle, stream=img, xref=img_xref)
    ##############

    # Crear una nueva imagen con fondo blanco
    width = 400
    height = 400
    image = Image.new('RGB', (width, height), 'white')

    # Crear un objeto ImageDraw para dibujar en la imagen
    draw = ImageDraw.Draw(image)

    # Especificar la fuente a utilizar
    font = ImageFont.truetype('Arial Unicode.ttf', size=40)

    # Especificar el texto y su posición en la imagen
    text_position = (10, 50)

    # Dibujar el texto en la imagen
    draw.text(text_position, datosFirmante, font=font, fill='black')

    # Guardar la imagen resultante
    image.save('imagen_con_texto.png')
    ##############
    # Open and flip the text image vertically
    text_image_path = 'imagen_con_texto.png'
    text_image = Image.open(text_image_path).transpose(Image.FLIP_TOP_BOTTOM)
    text_image.save('flipped_text_image.png')

    img1 = open('imagen_con_texto.png', "rb").read()  # an image file
    first_page.insert_image(fitz.Rect(250, 0, 450, 220), stream=img1, xref=img_xref)

    # Save the modified PDF
    file_handle.save(output_file)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def prueba_verificar(request):
    """
    ESte metodo sirve para actualizar la cola aws
    @type request: no recibe nada
    @rtype: no devuelve nada
    """
    env = environ.Env()
    environ.Env.read_env()  # LEE ARCHIVO .ENV
    client_s3 = boto3.client(
        's3',
        aws_access_key_id=env.str('AWS_ACCESS_KEY_ID'),
        aws_secret_access_key=env.str('AWS_SECRET_ACCESS_KEY')
    )
    with tempfile.TemporaryDirectory() as d:
        ruta = d + 'creditosPreAprobados.xlsx'
        s3 = boto3.resource('s3')
        s3.meta.client.download_file(env.str('AWS_STORAGE_BUCKET_NAME'),
                                     str('CORP/documentosCreditosPersonas/64906cb5b5913ced050fb5b2_solicitudCreditoFirmado.pdf'),
                                     ruta)

    input_file = "/Users/papamacone/Documents/Edgar/grp-back-coop/GlobalRedPyme/apps/CORP/corp_creditoPersonas/comandancia.pdf"
    output_file = "example-with-barcode.pdf"
    barcode_file = "/Users/papamacone/Documents/Edgar/grp-back-coop/GlobalRedPyme/output.png"

    # define the position (upper-right corner)
    image_rectangle = fitz.Rect(450, 20, 550, 120)

    # retrieve the first page of the PDF
    file_handle = fitz.open(input_file)
    first_page = file_handle[0]

    img = open(barcode_file, "rb").read()  # an image file
    img_xref = 0

    first_page.insert_image(image_rectangle, stream=img, xref=img_xref)

    file_handle.save(output_file)

    return Response({'verificada': 'hashok'}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
# @permission_classes([IsAuthenticated])
def verificarPropietarioFirma(request):
    """
    Este metodo sirve para verificar la firma electronica
    @type request: recibe el archivo, clavefirma, rucempresa
    @rtype: devuelve el mensaje de si verificada, caso contrario devuelve el error generado
    """
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'verificarPropietarioFirma',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'CREAR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        firmaCorresponde = usuarioPropietarioFirma(request.data['certificado'], request.data['claveFirma'],
                                                   request.data['rucEmpresa'])
        if firmaCorresponde is False:
            return Response({"message": "Lo sentimos, parece que la firma registrada no corresponde a "
                                        "la del representante legal o el dueño de la empresa o negocio. "
                                        "Por favor verifica la información y vuelve a intentarlo."},
                            status=status.HTTP_200_OK)

        return Response(status=status.HTTP_200_OK)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def recargar_lineas_creditos(request):
    """
    Este metodo sirve para subir el archivo de recargar lineas credito
    @type request: recibe el archivo excel
    @rtype: devuelve el registro creado, caso contrario devuelve el error generado
    """
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'recargar/lineasCreditos',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'CREAR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        contValidos = 0
        contInvalidos = 0
        contTotal = 0
        errores = []
        try:
            if request.method == 'POST':
                print('llegA')
                excel = request.FILES['documento']
                ruta = os.path.join(BASE_DIR, excel.name)
                # Leer los datos del archivo
                with open(ruta, 'wb') as archivo_destino:
                    for chunk in excel.chunks():
                        archivo_destino.write(chunk)
                print('paso')
                first = True  # si tiene encabezado
                #             uploaded_file = request.FILES['documento']
                # you may put validations here to check extension or file size
                wb = openpyxl.load_workbook(ruta)
                # getting a particular sheet by name out of many sheets
                worksheet = wb["Recargas"]
                lines = list()
            for row in worksheet.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                lines.append(row_data)

            for dato in lines:
                contTotal += 1
                if first:
                    first = False
                    continue
                else:
                    if len(dato) == 7:
                        resultadoInsertar = insertarDato_creditoPreaprobadoNegocio(dato,
                                                                                   request.data['empresa_financiera'])
                        if resultadoInsertar != 'Dato insertado correctamente':
                            contInvalidos += 1
                            errores.append(
                                {"error": "Error en la línea " + str(contTotal) + ": " + str(resultadoInsertar)})
                        else:
                            contValidos += 1
                    else:
                        contInvalidos += 1
                        errores.append({"error": "Error en la línea " + str(
                            contTotal) + ": la fila tiene un tamaño incorrecto (" + str(len(dato)) + ")"})

            result = {"mensaje": "La Importación se Realizo Correctamente",
                      "correctos": contValidos,
                      "incorrectos": contInvalidos,
                      "errores": errores
                      }
            os.remove(ruta)
            return Response(result, status=status.HTTP_201_CREATED)

        except Exception as e:
            err = {"error": 'Error verifique el archivo, un error ha ocurrido: {}'.format(e)}
            return Response(err, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


def insertarDato_creditoPreaprobadoNegocio(dato, empresa_financiera):
    """
    Este metodo sirve para registra en la tabla recargacredito
    @type empresa_financiera: recibe el idde la empresa financiera
    @type dato: recibe la fila del excel
    @rtype: no devuelve nada
    """
    try:
        info = {
            "nombreNegocio": dato[1],
            "rucNegocio": dato[2],
            "nombresRepresentante": dato[3],
            "rucRepresentante": dato[4],
            "correoRepresentante": dato[5],
        }
        timezone_now = timezone.localtime(timezone.now())
        data = {}
        data['fechaRecarga'] = dato[0].replace('"', "")[0:10] if dato[0] != "NULL" else None
        data['info'] = info
        data['monto'] = dato[6].replace('"', "") if dato[6] != "NULL" else None
        data['estado'] = 'Nuevo'
        data['tipoCredito'] = 'LineaCredito'
        data['observaciones'] = ''

        data['empresaIfis_id'] = empresa_financiera
        data['created_at'] = str(timezone_now)
        # inserto el dato con los campos requeridos
        RegarcarCreditos.objects.create(**data)
        return 'Dato insertado correctamente'
    except Exception as e:
        return str(e)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def listar_recargar_lineas_creditos(request):
    """
    Este metodo sirve para subir el archivo de recargar lineas credito
    @type request: recibe el archivo excel
    @rtype: devuelve el registro creado, caso contrario devuelve el error generado
    """
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'listar/recargar/lineasCreditos',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)
            # paginacion
            page_size = int(request.data['page_size'])
            page = int(request.data['page'])
            offset = page_size * page
            limit = offset + page_size
            # Filtros
            filters = {"state": "1"}

            if "empresaIfis_id" in request.data:
                if request.data["empresaIfis_id"] != '':
                    filters['empresaIfis_id'] = ObjectId(request.data["empresaIfis_id"])

            if "estado" in request.data:
                if request.data["estado"] != '':
                    filters['estado__icontains'] = request.data["estado"]

            if "tipoCredito" in request.data:
                if request.data["tipoCredito"] != '':
                    filters['tipoCredito'] = str(request.data["tipoCredito"])

            # Serializar los datos
            query = RegarcarCreditos.objects.filter(**filters).order_by('-created_at')
            serializer = RegarcarCreditosSerializer(query[offset:limit], many=True)
            new_serializer_data = {'cont': query.count(),
                                   'info': serializer.data}
            # envio de datos
            return Response(new_serializer_data, status=status.HTTP_200_OK)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def actualizar_recargar_lineas_creditos(request, pk):
    """
    Este metodo sirve para actualizar la recargar lineas credito
    @type request: recibe los campos de la tabla recarga creditos
    @rtype: devuelve el registro actualizado, caso contrario devuelve el error generado
    """
    request.POST._mutable = True
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'actualizar/recargar/lineasCreditos/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'ESCRIBIR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            logModel['dataEnviada'] = str(request.data)
            query = RegarcarCreditos.objects.filter(pk=ObjectId(pk), state=1).order_by('-created_at').first()
        except RegarcarCreditos.DoesNotExist:
            errorNoExiste = {'error': 'No existe'}
            createLog(logModel, errorNoExiste, logExcepcion)
            return Response(status=status.HTTP_404_NOT_FOUND)
        if request.method == 'POST':
            now = timezone.localtime(timezone.now())
            request.data['updated_at'] = str(now)
            if 'created_at' in request.data:
                request.data.pop('created_at')

            serializer = RegarcarCreditosSerializer(query, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                if 'estado' in request.data and 'Aprobado' == request.data['estado']:
                    enviarCorreoRecarga(serializer.data)
                    # Vamos a registrar la transacction del credito
                    credito = CreditoPersonas.objects.filter(rucEmpresa=serializer.data['info']['rucRepresentante'],
                                                             estado='Aprobado',
                                                             state=1).order_by('-created_at').first()
                    Transacciones.objects.create(**{
                        'fechaTransaccion': timezone_now,
                        'tipo': 'Recarga de cupo de línea de crédito',
                        'estado': 'Aprobado',
                        'informacion': json.dumps(serializer.data),
                        'egreso': float(serializer.data['monto']),
                        'total': (credito.montoDisponible + float(serializer.data['monto'])),
                        'user_id': credito.user_id,
                        'creditoPersona_id': credito._id,
                        'regarcarCreditos': query,
                    })
                return Response(serializer.data)
            createLog(logModel, serializer.errors, logExcepcion)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


def enviarCorreoRecarga(data):
    """
    Este metodo sirve para enviar el correo de credito negado
    @type email: recibe el email
    @type montoAprobado: recibe el monto aprobado
    @rtype: no devuelve nada
    """
    empresaIfis = Empresas.objects.filter(tipoEmpresa='ifis', state=1).order_by('-created_at').first()
    subject, from_email, to = 'RECARGA DE CUPO EN LÍNEA DE CRÉDITO', "08d77fe1da-d09822@inbox.mailtrap.io", \
                              data['info']['correoRepresentante']
    txt_content = f"""
        Estimad@ {data['info']['nombresRepresentante']}

        La Cooperativa {empresaIfis.nombreEmpresa} ha recargado el Monto de su línea de crédito con {data['monto']}

        Atentamente,

        {empresaIfis.nombreEmpresa}
    """
    html_content = f"""
        <html>
            <body>
                <h1>
                Estimad@ {data['info']['nombresRepresentante']}
                </h1>
                <br>
                <p>
                La Cooperativa {empresaIfis.nombreEmpresa} ha recargado el Monto de su línea de crédito con {data['monto']}
                </p>
                <br>
                Atentamente,
                <br>
                {empresaIfis.nombreEmpresa}
            </body>
        </html>
    """
    sendEmail(subject, txt_content, from_email, to, html_content)


def enviarCorreoNegarLineaCredito(email, motivo):
    """
    Este metodo sirve para enviar el correo de credito negado
    @type email: recibe el email
    @type montoAprobado: recibe el monto aprobado
    @rtype: no devuelve nada
    """
    empresaIfis = Empresas.objects.filter(tipoEmpresa='ifis', state=1).order_by('-created_at').first()
    subject, from_email, to = 'NEGAR LÍNEA DE CRÉDITO', "08d77fe1da-d09822@inbox.mailtrap.io", email
    txt_content = f"""
        Estimad@

        La Cooperativa {empresaIfis.nombreEmpresa} ha negado la activación por el motivo {motivo}

        Atentamente,

        {empresaIfis.nombreEmpresa}
    """
    html_content = f"""
        <html>
            <body>
                <h1>
                Estimad@
                </h1>
                <br>
                <p>
                La Cooperativa {empresaIfis.nombreEmpresa} ha negado la activación por el motivo {motivo}
                </p>
                <br>
                Atentamente,
                <br>
                {empresaIfis.nombreEmpresa}
            </body>
        </html>
    """
    sendEmail(subject, txt_content, from_email, to, html_content)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def creditoPersonas_create_local(request):
    """
    Este metod sirve para crear un creditopersona
    @type request: recibe los campos de la tabla credito persona
    @rtype: DEvuelve el registro creado, caso contrario devuelve el error generado
    """
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'create/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'CREAR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)
            request.data['created_at'] = str(timezone_now)
            if 'updated_at' in request.data:
                request.data.pop('updated_at')

            if 'nombres' in request.data:
                if request.data['nombres'] != "":
                    request.data['nombresCompleto'] = request.data['nombres'] + ' ' + request.data['apellidos']

            serializer = CreditoPersonasSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                if 'tipoCredito' in serializer.data and serializer.data['tipoCredito'] == 'null':
                    usuario = serializer.data['user']
                    nombreGarante = usuario['garante']['nombres'] + ' ' + usuario['garante']['apellidos']
                    nombreSolicitante = usuario['nombres'] + ' ' + usuario['apellidos']
                    enviarCorreoSolicitudGarante(usuario['garante']['correoGarante'], serializer.data['_id'],
                                                 nombreGarante, nombreSolicitante)
                createLog(logModel, serializer.data, logTransaccion)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            createLog(logModel, serializer.errors, logExcepcion)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)


def enviarCorreoSolicitudGarante(email, id, garante, solicitante):
    """
    Este metodo sirve para enviar el correo de credito negado
    @type email: recibe el email
    @type montoAprobado: recibe el monto aprobado
    @rtype: no devuelve nada
    """
    subject, from_email, to = 'Autorización de la Madrina/Padrino', "credicompra.bigpuntos@corporacionomniglobal.com", \
                              email
    txt_content = f"""
        Autorización de Crédito de Consumo

        Estimad@ {garante}, {solicitante} desea que usted le apadrine para poder acceder a un Crédito de Consumo para realizar compras.

        Para confirmar su aprobación como Garantía para acceder al Crédito, haga click en el siguiente enlace y confirme sus datos: 
        {config.API_FRONT_END_CENTRAL}/pages/confirmacion-garante/{id}

        Atentamente,
        CrediCompra – Big Puntos
    """
    html_content = f"""
        <html>
            <body>
                <h1>
                Autorización de Crédito de Consumo
                </h1>
                <p>
                Estimad@ {garante}, {solicitante} desea que usted le apadrine para poder acceder a un Crédito
                 de Consumo para realizar compras.
                </p>
                <br>
                <p>
                Para confirmar su aprobación como Garantía para acceder al Crédito, haga click en el siguiente enlace
                 y confirme sus datos: <a href='{config.API_FRONT_END_CENTRAL}/pages/confirmacion-garante/{id}'>ENLACE</a>
                </p>
                <br>
                Atentamente,
                <br>
                CrediCompra – Big Puntos
                <br>
            </body>
        </html>
    """
    sendEmail(subject, txt_content, from_email, to, html_content)


# ACTUALIZAR SIN AUTENTICAR
@api_view(['POST'])
def creditoPersonas_update_sinAutenticar(request, pk):
    """
    Este metod sirve para actualizar un creditopersona
    @type pk: recibe el id de la tabla credito persona
    @type request: recibe los campos de la tabla credito personas
    @rtype: DEvuelve el registro obtenido, caso contrario devuelve el error generado
    """
    request.POST._mutable = True
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'update/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'ESCRIBIR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            logModel['dataEnviada'] = str(request.data)
            query = CreditoPersonas.objects.filter(pk=ObjectId(pk), state=1).first()
        except CreditoPersonas.DoesNotExist:
            errorNoExiste = {'error': 'No existe'}
            createLog(logModel, errorNoExiste, logExcepcion)
            return Response(status=status.HTTP_404_NOT_FOUND)
        if request.method == 'POST':
            now = timezone.localtime(timezone.now())
            request.data['updated_at'] = str(now)
            if 'created_at' in request.data:
                request.data.pop('created_at')
            serializer = CreditoPersonasSerializer(query, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                createLog(logModel, serializer.data, logTransaccion)
                return Response(serializer.data)
            createLog(logModel, serializer.errors, logExcepcion)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)
