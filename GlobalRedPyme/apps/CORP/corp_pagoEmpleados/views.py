import io
import qrcode
import fitz
from cryptography.hazmat import backends
from cryptography.hazmat.primitives.serialization import pkcs12
from endesive.pdf import cms
## Libreria para agregar imagenes a pdf
from PIL import Image, ImageDraw, ImageFont

from django.core.files.uploadedfile import InMemoryUploadedFile
from .models import PagoEmpleados
from .serializers import (
    PagoEmpleadosSerializer
)
from rest_framework import status
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from django.utils import timezone
import json
# excel
import openpyxl

from ..corp_creditoPersonas.models import CreditoPersonas
from ..corp_movimientoCobros.models import Transacciones
# Utils
from ...utils import utils
# Import PDF
from fpdf import FPDF
# Enviar Correo
from apps.config.util import sendEmail
# ObjectId
from bson import ObjectId
# logs
from ...CENTRAL.central_logs.methods import createLog, datosTipoLog, datosProductosMDP

# declaracion variables log
datosAux = datosProductosMDP()
datosTipoLogAux = datosTipoLog()
# asignacion datos modulo
logModulo = datosAux['modulo']
logApi = datosAux['api']
# asignacion tipo de datos
logTransaccion = datosTipoLogAux['transaccion']
logExcepcion = datosTipoLogAux['excepcion']


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def uploadEXCEL_pagosEmpleados(request):
    """
    ESte metodo sirve para cargar los pagos a los empleados
    @type request: El campo request recibe el archivo excel
    @rtype: DEvuele la lista de los ingresos correctos, incorrectos, caso contrario devuelve el error generado
    """
    contValidos = 0
    contInvalidos = 0
    contTotal = 0
    errores = []
    try:
        if request.method == 'POST':
            first = True  # si tiene encabezado
            uploaded_file = request.FILES['documento']
            # you may put validations here to check extension or file size
            wb = openpyxl.load_workbook(uploaded_file)
            # getting a particular sheet by name out of many sheets
            worksheet = wb["Pagos"]
            lines = list()
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            lines.append(row_data)

        for dato in lines:
            contTotal += 1
            if first:
                first = False
                continue
            else:
                if len(dato) == 12:
                    resultadoInsertar = insertarDato_PagoEmpleado(dato, request.data['user_id'])
                    if resultadoInsertar != 'Dato insertado correctamente':
                        contInvalidos += 1
                        errores.append({"error": "Error en la línea " + str(contTotal) + ": " + str(resultadoInsertar)})
                    else:
                        contValidos += 1
                else:
                    contInvalidos += 1
                    errores.append({"error": "Error en la línea " + str(
                        contTotal) + ": la fila tiene un tamaño incorrecto (" + str(len(dato)) + ")"})

        result = {"mensaje": "La Importación se Realizo Correctamente",
                  "correctos": contValidos,
                  "incorrectos": contInvalidos,
                  "errores": errores
                  }
        return Response(result, status=status.HTTP_201_CREATED)

    except Exception as e:
        err = {"error": 'Error verifique el archivo, un error ha ocurrido: {}'.format(e)}
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def pagoEmpleados_update(request, pk):
    """
    El metodo sirve para actualizar el pago del empleado
    @type pk: El campo pk recibe el id del pago empleado
    @type request: El campo request recibe los campos del pago empleado
    @rtype: DEvuelve el registro actualizado, caso contrario devuelve el error generado
    """
    request.POST._mutable = True
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'update/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'ESCRIBIR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            logModel['dataEnviada'] = str(request.data)
            query = PagoEmpleados.objects.filter(pk=ObjectId(pk), state=1).first()
        except PagoEmpleados.DoesNotExist:
            errorNoExiste = {'error': 'No existe'}
            createLog(logModel, errorNoExiste, logExcepcion)
            return Response(status=status.HTTP_404_NOT_FOUND)
        if request.method == 'POST':
            now = timezone.localtime(timezone.now())
            request.data['updated_at'] = str(now)
            if 'created_at' in request.data:
                request.data.pop('created_at')

            if 'claveFirma' in request.data:
                if request.data != '':
                    datau, datas = firmar(request)
                    archivo_pdf_para_enviar_al_cliente = io.BytesIO()
                    archivo_pdf_para_enviar_al_cliente.write(datau)
                    archivo_pdf_para_enviar_al_cliente.write(datas)
                    archivo_pdf_para_enviar_al_cliente.seek(0)

                    request.data['archivoFirmado'] = InMemoryUploadedFile(archivo_pdf_para_enviar_al_cliente,
                                                                          'media',
                                                                          'documentoFirmado.pdf',
                                                                          'application/pdf',
                                                                          archivo_pdf_para_enviar_al_cliente.getbuffer().nbytes,
                                                                          None
                                                                          )
                    request.data['fechaFirma'] = str(now)

            serializer = PagoEmpleadosSerializer(query, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()

                if serializer.data['estado'] == 'Negado':
                    registro = serializer.data
                    envioCorreoNegado(registro['correo'], registro['nombresCompletos'], registro['observacion'],
                                      registro['montoPagar'])
                if serializer.data['estado'] == 'Aprobado':
                    registro = serializer.data
                    envioCorreoAprobado(registro['empresa']['correo'], registro['nombresCompletos'],
                                        registro['montoPagar'], registro['montoDisponible'], registro)
                    nombrePyme = registro['empresa']['comercial']
                    nombreReresentanteLegal = registro['empresa']['reprsentante']
                    envioCorreoTranserencia(registro['correo'], registro['montoPagar'], registro['nombresCompletos'],
                                            nombreReresentanteLegal, nombrePyme, registro['mesPago'])
                    # Vamos a registrar la transacction del credito
                    credito = CreditoPersonas.objects.filter(user_id=serializer.data['user_id'], estado='Aprobado',
                                                             state=1).order_by('-created_at').first()
                    Transacciones.objects.create(**{
                        'fechaTransaccion': timezone_now,
                        'tipo': 'Pago empleado',
                        'estado': 'Aprobado',
                        'informacion': json.dumps(serializer.data),
                        'egreso': float(serializer.data['montoPagar']),
                        'total': (credito.montoDisponible - float(serializer.data['montoPagar'])),
                        'user_id': serializer.data['user_id'],
                        'creditoPersona_id': credito._id,
                        'pagoEmpleados': query,
                    })

                createLog(logModel, serializer.data, logTransaccion)
                return Response(serializer.data)
            createLog(logModel, serializer.errors, logExcepcion)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def pagoEmpleados_list(request):
    """
    ESte metodo sirve para listar los pagos de los empleados
    @type request: El campo request recibe page, page_size, user_id, estado
    @rtype: DEVUELVE una lista de pagos empleados, caso contrario devuelve el error generado
    """
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'list/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)
            # paginacion
            page_size = int(request.data['page_size'])
            page = int(request.data['page'])
            offset = page_size * page
            limit = offset + page_size
            # Filtros
            filters = {"state": "1"}

            if "user_id" in request.data and request.data["user_id"]:
                filters['user_id'] = str(request.data["user_id"])

            if "estado" in request.data and request.data["estado"]:
                filters['estado__in'] = request.data["estado"]

            # Serializar los datos
            query = PagoEmpleados.objects.filter(**filters).order_by('-created_at')
            serializer = PagoEmpleadosSerializer(query[offset:limit], many=True)
            new_serializer_data = {'cont': query.count(),
                                   'info': serializer.data}
            # envio de datos
            return Response(new_serializer_data, status=status.HTTP_200_OK)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)


from cryptography.hazmat import backends
from cryptography.hazmat.primitives.serialization import pkcs12
from endesive.pdf import cms


def firmar(request):
    """
    Este metodo sirve para firma un documento
    @type request: El campo request recibe el documento que se va firmar
    @rtype: Devuelve el documento firmado
    """
    usuario = json.loads(request.data['usuarioEmpresa'])
    certificado = request.data['certificado']
    pdf = request.data['archivoFirmado']
    contrasenia = request.data['claveFirma']
    date = timezone_now = timezone.localtime(timezone.now())
    date = date.strftime("D:%Y%m%d%H%M%S+00'00'")
    dct = {
        "aligned": 0,
        "sigflags": 3,
        "sigflagsft": 132,
        "sigpage": 0,
        "sigbutton": True,
        "sigfield": "Signature1",
        "auto_sigfield": True,
        "sigandcertify": True,
        "signaturebox": (470, 840, 570, 640),
        "signature": usuario['nombresCompleto'],
        "signature_manual": [],
        # "signature_img": "signature_test.png",
        "contact": usuario['email'],
        "location": "Ubicación",
        "signingdate": date,
        "reason": "Pago a empleado",
        "password": contrasenia,
    }
    # with open("cert.p12", "rb") as fp:
    p12 = pkcs12.load_key_and_certificates(
        certificado.read(), contrasenia.encode("ascii"), backends.default_backend()
    )
    datosFirmante = f"""FIRMADO POR:\n {dct['signature']} \n FECHA:\n {date}"""
    generarQR(datosFirmante)
    output_file = "example-with-barcode.pdf"
    agregarQRDatosFirmante(datosFirmante, output_file, pdf)

    # datau = open(fname, "rb").read()
    datau = open(output_file, "rb").read()
    dct.pop('signature')
    datas = cms.sign(datau, dct, p12[0], p12[1], p12[2], "sha256")
    return datau, datas
    # return Response('new_serializer_data', status=status.HTTP_200_OK)


def enviarNegadoPago(email, monto):
    """
    Este metodo sirve para enviar el correo de pago negado
    @type monto: El campo monto recibe la cantidad de monto
    @type email: El campo email recibe el email del usuario
    @rtype: No devuelve nada
    """
    subject, from_email, to = 'RAZÓN POR LA QUE SE NIEGA EL PAGO A PROVEEDORES', "08d77fe1da-d09822@inbox.mailtrap.io", \
                              email
    txt_content = f"""
        PAGO A PROVEEDORES - CRÉDITO PAGOS

        Lo sentimos!!

        La transferencia por ${monto} DE LA FACTURA A PAGAR ha sido rechazada. 
        Por favor revise sus fondos e intente de nuevo.

        Si cree que es un error, contáctese con su agente a través de https://walink.co/b5e9c0

        Atentamente,

        Global RedPyme – Crédito Pagos
    """
    html_content = f"""
                <html>
                    <body>
                        <h1>PAGO A PROVEEDORES - CRÉDITO PAGOS</h1>
                        <br>
                        <h3><b>Lo sentimos!!</b></h3>
                        <br>
                        <p>La transferencia por ${monto} DE LA FACTURA A PAGAR ha sido rechazada. 
                        Por favor revise sus fondos e intente de nuevo.</p>
                        <br>
                        <br>
                        <p>Si cree que es un error, contáctese con su agente a través de https://walink.co/b5e9c0</p>
                        <br>
                        Atentamente,
                        <br>
                        Global RedPyme – Crédito Pagos
                        <br>
                    </body>
                </html>
                """
    sendEmail(subject, txt_content, from_email, to, html_content)


def enviarProcesandoPago(email, monto):
    """
    Este metodo sirve para enviar el correo de pago procesado
    @type monto: El campo monto recibe la cantidad de monto
    @type email: El campo email recibe el email del usuario
    @rtype: No devuelve nada
    """
    subject, from_email, to = 'Transferencia exitosa', "08d77fe1da-d09822@inbox.mailtrap.io", \
                              email
    txt_content = f"""
        PAGO A PROVEEDORES - CRÉDITO PAGOS

        FELICIDADES!!

        La transferencia por ${monto} ha sido realizada con éxito, 
        adjuntamos el comprobante del pago realizado. 
        En 24 horas será acreditado a la cuenta destino.

        Atentamente,

        Global RedPyme – Crédito Pagos
    """
    html_content = f"""
                <html>
                    <body>
                        <h1>PAGO A PROVEEDORES - CRÉDITO PAGOS</h1>
                        <br>
                        <h3><b>FELICIDADES!!</b></h3>
                        <br>
                        <p>La transferencia por ${monto} ha sido realizada con éxito, 
                        adjuntamos el comprobante del pago realizado. 
                        En 24 horas será acreditado a la cuenta destino.</p>
                        <br>
                        <br>
                        Atentamente,
                        <br>
                        Global RedPyme – Crédito Pagos
                        <br>
                    </body>
                </html>
                """
    sendEmail(subject, txt_content, from_email, to, html_content)


def generarPDF(datos):
    """
    ESte metodo sirve para generar un pdf
    @type datos: El campo datos no recibe nada
    @rtype: Devuelve el docmuento pdf
    """
    # save FPDF() class into a
    # variable pdf
    pdf = FPDF()

    # Add a page
    pdf.add_page()

    # set style and size of font
    # that you want in the pdf
    pdf.set_font("Arial", size=15)

    # create a cell
    pdf.cell(200, 10, txt="GeeksforGeeks",
             ln=1, align='C')

    # add another cell
    pdf.cell(200, 10, txt="A Computer Science portal for geeks.",
             ln=2, align='C')

    return pdf


def insertarDato_PagoEmpleado(dato, user_id):
    """
    ESte metodo sirve para gardar en la tabla de pago empleado
    @type user_id: El campo user_id recibe el id del usuario
    @type dato: El campo dato recibe la fila del excel
    @rtype: DEvuelve el registro guardado, caso contrario devuelve el error generado
    """
    try:
        if ('None' in dato):
            return 'Tiene campos vacios'
        if (not utils.__validar_ced_ruc(dato[5], 0)):
            return 'Cedula incorrecta'
        if (not utils.isValidEmail(dato[9])):
            return 'Email incorrecto'
        if (not utils.isValidTelefono(dato[8])):
            return 'Celular incorrecto'
        if (not utils.isValidTelefono(dato[8])):
            return 'Whatsapp incorrecto'
        data = {}
        data['nombresCompletos'] = f"{dato[6]} {dato[7]}"
        data['cedula'] = dato[5]
        data['celular'] = dato[8]
        data['correo'] = dato[9]
        data['montoPagar'] = dato[3]
        data['codigoEmpleado'] = dato[4]
        data['mesPago'] = dato[1]
        data['anio'] = dato[2]
        data['numeroCuentaEmpleado'] = dato[10]
        data['bancoDestino'] = dato[11]
        data['estado'] = ''
        data['user_id'] = user_id
        data['state'] = 1
        # inserto el dato con los campos requeridos
        PagoEmpleados.objects.update_or_create(**data)
        return 'Dato insertado correctamente'
    except Exception as e:
        return str(e)


def envioCorreoNegado(email, nombresCompletosEmpleado, observacion, montoPagar):
    """
    Este metodo sirve para enviar el correo negado
    @type montoPagar: El campo montopagar recibe el monto a pagar
    @type observacion: El campo observacion recibe la observacion
    @type nombresCompletosEmpleado: El campo recibe los nombres del usuario
    @type email: El campo email recibe el email del usuario
    @rtype: No devuelve nada
    """
    subject, from_email, to = 'PAGO FALLIDO', "08d77fe1da-d09822@inbox.mailtrap.io", \
                              email
    txt_content = f"""
            Pago a empleados – Crédito Pagos

            Lo sentimos

            Su pago a empleados por ${montoPagar} ha sido NEGADO debido a {observacion}

            Si necesita ayuda personalizada, contáctese con un asesor a través del siguiente enlace: https://wa.link/koof8g 

            Atentamente,

            Global RedPyme – Crédito Pagos
        """
    html_content = f"""
        <html>
            <body>
                <h1>Pago a empleados – Crédito Pagos</h1>
                <br>
                <h3><b>Lo sentimos</b></h3>
                <br>
                <p>Su pago a empleados por ${montoPagar} ha sido NEGADO debido a {observacion}</p>
                <br>
                <br>
                <p>Si necesita ayuda personalizada, contáctese con un asesor a través del siguiente enlace: https://wa.link/koof8g </p>
                <br>
                Atentamente,
                <br>
                Cooperativa San José de Vittoria – Crédito Pagos
                <br>
            </body>
        </html>
    """
    sendEmail(subject, txt_content, from_email, to, html_content)


def envioCorreoAprobado(email, nombresCompletosEmpleado, monto, montoDisponible, registro):
    """
    ESte metodo sirve para enviar el correo aprobado
    @param registro: Este campo recibe el registro guardado
    @type montoDisponible: El campo recibe el monto disponible
    @type monto: El campo recibe el monto
    @type nombresCompletosEmpleado: El campo recibe el nombre del usuario
    @type email: El campo recibe el email
    @rtype: No devuelve nada
    """
    subject, from_email, to = 'Transferencia existosa', "08d77fe1da-d09822@inbox.mailtrap.io", \
                              email
    txt_content = f"""
            Pago a empleados – Crédito Pagos

            FELICIDADES!!

            Usted acaba de realizar un pago a su empleado por ${monto}. A continuación le mostramos un resumen de su pago:

            Número de comprobante de transferencia: {registro['numeroComprobante']}
            Fecha de transferencia: {registro['fechaProceso']}
            Nombre del empleado: {registro['nombresCompletos']}
            Numero de cédula del empleado: {registro['cedula']}
            Monto pagado: {monto}
            Número de cuenta del empleado: {registro['numeroCuentaEmpleado']}
            Banco destino: {registro['bancoDestino']}
            Estado: APROBADO

            Atentamente,

            Cooperativa San José de Vittoria – Crédito Pagos
        """
    html_content = f"""
        <html>
            <body>
                <h1>Pago a empleados – Crédito Pagos</h1>
                <br>
                <h3><b>FELICIDADES!!</b></h3>
                <br>
                <p>Usted acaba de realizar un pago a su empleado por ${monto}. A continuación le mostramos un resumen de su pago:</p>
                <br>
                Número de comprobante de transferencia: {registro['numeroComprobante']}
                Fecha de transferencia: {registro['fechaProceso']}
                Nombre del empleado: {registro['nombresCompletos']}
                Numero de cédula del empleado: {registro['cedula']}
                Monto pagado: {monto}
                Número de cuenta del empleado: {registro['numeroCuentaEmpleado']}
                Banco destino: {registro['bancoDestino']}
                Estado: APROBADO
                <br>
                Atentamente,
                <br>
                Cooperativa San José de Vittoria – Crédito Pagos
                <br>
            </body>
        </html>
    """
    sendEmail(subject, txt_content, from_email, to, html_content)


def envioCorreoTranserencia(email, monto, nombresCompletosEmpleado, nombreRepresentanteLegal, nombrePyme, mesPago):
    """
    ESte metodo sirve para enviar el correo aprobado
    @param mesPago: Recibe el mes de pago
    @param nombrePyme: REcibe el nombre de la pyme
    @param nombreRepresentanteLegal: El campo recibe el nombre del representante
    @type monto: El campo recibe el monto
    @type nombresCompletosEmpleado: El campo recibe el nombre del usuario
    @type email: El campo recibe el email
    @rtype: No devuelve nada
    """
    subject, from_email, to = 'TRANSFERENCIA RECIBIDA', "08d77fe1da-d09822@inbox.mailtrap.io", \
                              email
    txt_content = f"""
            TRANSFERENCIA RECIBIDA

            Usted ha recibido una transferencia por parte de {nombreRepresentanteLegal},

            Estimado/o {nombresCompletosEmpleado} usted ha recibido una transferencia en su cuenta bancaria por un 
            monto de {monto} por sus labores en la empresa {nombrePyme} respecto al mes de {mesPago}

            Atentamente,

            Equipo Global Redpyme – Crédito Pagos
        """
    html_content = f"""
        <html>
            <body>
                <h1>TRANSFERENCIA RECIBIDA</h1>
                <br>
                <h3><b>Usted ha recibido una transferencia por parte de {nombreRepresentanteLegal},</b></h3>
                <br>
                <p>Estimado/o {nombresCompletosEmpleado} usted ha recibido una transferencia en su cuenta bancaria por
                 un monto de {monto} por sus labores en la empresa {nombrePyme}  respecto al mes de {mesPago}
                 </p>
                <br>
                <br>
                Atentamente,
                <br>
                Equipo Global Redpyme – Crédito Pagos
                <br>
            </body>
        </html>
    """
    sendEmail(subject, txt_content, from_email, to, html_content)


def generarQR(datos):
    """
    ESte metodo sirve para generar un QR
    @type datos: Recibe la informacion del usuario
    @rtype: No devuelve nada
    """
    img = qrcode.make(datos)
    f = open("output.png", "wb")
    img.save(f)
    f.close()


def agregarQRDatosFirmante(datosFirmante, output_file, ruta):
    """
    ESte metodo sirve para agregar el qr a un documento pdf
    @param ruta: recibe la ruta de donde este el qr
    @param output_file: recibe la ruta donde se va guardar el documento generado
    @type datosFirmante: REcibe los datos del usuario que firma
    @rtype: no devuelve nada
    """
    # Define the position and size of the image rectangle
    image_rectangle = fitz.Rect(0, 420, 150, 520)  # Adjust the coordinates and size as needed

    # Retrieve the first page of the PDF
    with open('temp_file.pdf', 'wb') as temp_file:
        temp_file.write(ruta.read())

    # 2. Abrir el archivo temporal con fitz.open()
    file_handle = fitz.open('temp_file.pdf')
    first_page = file_handle[0]

    # Open and flip the image vertically
    image_path = 'output.png'
    image = Image.open(image_path).transpose(Image.FLIP_TOP_BOTTOM)
    image.save('flipped_image.png')

    # Insert the flipped image into the PDF
    img = open('output.png', "rb").read()  # an image file
    img_xref = 0
    first_page.insert_image(image_rectangle, stream=img, xref=img_xref)
    ##############

    # Crear una nueva imagen con fondo blanco
    width = 400
    height = 400
    image = Image.new('RGB', (width, height), 'white')

    # Crear un objeto ImageDraw para dibujar en la imagen
    draw = ImageDraw.Draw(image)

    # Especificar la fuente a utilizar
    font = ImageFont.truetype('Arial Unicode.ttf', size=40)

    # Especificar el texto y su posición en la imagen
    text_position = (10, 50)

    # Dibujar el texto en la imagen
    draw.text(text_position, datosFirmante, font=font, fill='black')

    # Guardar la imagen resultante
    image.save('imagen_con_texto.png')
    ##############
    # Open and flip the text image vertically
    text_image_path = 'imagen_con_texto.png'
    text_image = Image.open(text_image_path).transpose(Image.FLIP_TOP_BOTTOM)
    text_image.save('flipped_text_image.png')

    img1 = open('imagen_con_texto.png', "rb").read()  # an image file
    first_page.insert_image(fitz.Rect(150, 420, 300, 520), stream=img1, xref=img_xref)

    # Save the modified PDF
    file_handle.save(output_file)
